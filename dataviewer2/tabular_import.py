from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import numpy as np
from PySide6 import QtCore, QtGui, QtWidgets

from .config import ImportedPointLayerDefinition
from .models import PointLayerData


SYMBOLS = [
    "Circle", "Square", "Triangle", "Triangle Down", "Diamond",
    "Star", "Cross", "Plus", "Pentagon", "Hexagon",
]


def _unique_headers(values: list[Any]) -> list[str]:
    result: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ""
        base = base or f"Column {index}"
        key = base.casefold()
        used[key] = used.get(key, 0) + 1
        result.append(base if used[key] == 1 else f"{base} ({used[key]})")
    return result


def excel_sheet_names(path: str | Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel import requires openpyxl. Install it with: pip install openpyxl") from exc
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def table_headers(path: str | Path, sheet_name: str = "") -> list[str]:
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book[sheet_name] if sheet_name else book.active
            row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            return _unique_headers(list(row))
        finally:
            book.close()
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return _unique_headers(next(csv.reader(handle, dialect), []))


def _read_rows(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book[sheet_name] if sheet_name else book.active
            iterator = sheet.iter_rows(values_only=True)
            headers = _unique_headers(list(next(iterator, ())))
            return headers, [list(row) for row in iterator]
        finally:
            book.close()
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        iterator = csv.reader(handle, dialect)
        headers = _unique_headers(next(iterator, []))
        return headers, [list(row) for row in iterator]


def load_imported_point_data(
    config_dir: Path, definition: ImportedPointLayerDefinition
) -> PointLayerData:
    path = Path(definition.source_file)
    if not path.is_absolute():
        path = config_dir / path
    if not path.exists():
        raise FileNotFoundError(f"Imported layer source file was not found: {path}")
    headers, rows = _read_rows(path, definition.sheet_name)
    lookup = {name: index for index, name in enumerate(headers)}
    if definition.x_field not in lookup or definition.y_field not in lookup:
        raise ValueError("The saved X or Y coordinate column no longer exists in the file.")
    x_index, y_index = lookup[definition.x_field], lookup[definition.y_field]
    valid_rows: list[tuple[int, list[Any], float, float]] = []
    for source_index, row in enumerate(rows, start=2):
        try:
            x = float(row[x_index])
            y = float(row[y_index])
        except (IndexError, TypeError, ValueError):
            continue
        if np.isfinite(x) and np.isfinite(y):
            valid_rows.append((source_index, row, x, y))
    if not valid_rows:
        raise ValueError("No rows contain valid numeric values in both coordinate columns.")

    metadata: dict[str, np.ndarray] = {}
    for column_index, header in enumerate(headers):
        metadata[header] = np.asarray([
            row[column_index] if column_index < len(row) and row[column_index] is not None else ""
            for _, row, _, _ in valid_rows
        ], dtype=object)
    if definition.label_fields:
        label_indices = [lookup[name] for name in definition.label_fields if name in lookup]
        metadata["Label"] = np.asarray([
            definition.label_separator.join(
                str(row[index]).strip()
                for index in label_indices
                if index < len(row) and row[index] not in (None, "")
            )
            for _, row, _, _ in valid_rows
        ], dtype=object)
    return PointLayerData(
        name=definition.name,
        x=np.asarray([item[2] for item in valid_rows], dtype=np.float64),
        y=np.asarray([item[3] for item in valid_rows], dtype=np.float64),
        source_index=np.asarray([item[0] for item in valid_rows], dtype=np.int64),
        metadata=metadata,
    )


class TabularPointImportDialog(QtWidgets.QDialog):
    def __init__(self, path: str | Path, group_names: list[str], parent=None) -> None:
        super().__init__(parent)
        self.path = Path(path)
        self.color = QtGui.QColor("#ff80ab")
        self.setWindowTitle("Import Excel / CSV point layer")
        self.resize(590, 650)
        layout = QtWidgets.QVBoxLayout(self)
        form = QtWidgets.QFormLayout()
        self.file_label = QtWidgets.QLabel(str(self.path))
        self.file_label.setWordWrap(True)
        form.addRow("File:", self.file_label)
        self.sheet_combo = QtWidgets.QComboBox()
        if self.path.suffix.lower() in {".xlsx", ".xlsm"}:
            self.sheet_combo.addItems(excel_sheet_names(self.path))
            form.addRow("Worksheet:", self.sheet_combo)
            self.sheet_combo.currentTextChanged.connect(self._load_headers)
        self.name_edit = QtWidgets.QLineEdit(self.path.stem)
        self.group_combo = QtWidgets.QComboBox(); self.group_combo.setEditable(True)
        self.group_combo.addItems(group_names); self.group_combo.setCurrentText("Imported Layers")
        self.x_combo = QtWidgets.QComboBox(); self.y_combo = QtWidgets.QComboBox()
        form.addRow("Layer name:", self.name_edit)
        form.addRow("Group:", self.group_combo)
        form.addRow("X / Easting column:", self.x_combo)
        form.addRow("Y / Northing column:", self.y_combo)
        layout.addLayout(form)

        layout.addWidget(QtWidgets.QLabel("Label columns (check two or more to merge them):"))
        self.label_columns = QtWidgets.QListWidget()
        self.label_columns.setAlternatingRowColors(True)
        layout.addWidget(self.label_columns, 1)
        label_options = QtWidgets.QFormLayout()
        self.separator_edit = QtWidgets.QLineEdit(" ")
        self.separator_edit.setToolTip("Text inserted between merged column values")
        self.show_labels = QtWidgets.QCheckBox(); self.show_labels.setChecked(True)
        label_options.addRow("Label separator:", self.separator_edit)
        label_options.addRow("Show labels:", self.show_labels)
        layout.addLayout(label_options)

        style_form = QtWidgets.QFormLayout()
        self.color_button = QtWidgets.QPushButton()
        self.color_button.clicked.connect(self._choose_color)
        self.size_spin = QtWidgets.QDoubleSpinBox(); self.size_spin.setRange(1, 30); self.size_spin.setValue(7)
        self.symbol_combo = QtWidgets.QComboBox(); self.symbol_combo.addItems(SYMBOLS)
        self.save_layer = QtWidgets.QCheckBox("Copy the file into the project and reload this layer at startup")
        self.save_layer.setChecked(True)
        style_form.addRow("Point color:", self.color_button)
        style_form.addRow("Point size:", self.size_spin)
        style_form.addRow("Point symbol:", self.symbol_combo)
        layout.addLayout(style_form)
        layout.addWidget(self.save_layer)
        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._validate); buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self._update_color_button()
        self._load_headers()

    def _load_headers(self) -> None:
        headers = table_headers(self.path, self.sheet_combo.currentText())
        self.x_combo.clear(); self.y_combo.clear(); self.label_columns.clear()
        self.x_combo.addItems(headers); self.y_combo.addItems(headers)
        for index, header in enumerate(headers):
            item = QtWidgets.QListWidgetItem(header)
            item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.CheckState.Unchecked)
            self.label_columns.addItem(item)
            lower = header.casefold()
            if lower in {"x", "easting", "east", "longitude", "lon"}:
                self.x_combo.setCurrentIndex(index)
            if lower in {"y", "northing", "north", "latitude", "lat"}:
                self.y_combo.setCurrentIndex(index)

    def _choose_color(self) -> None:
        value = QtWidgets.QColorDialog.getColor(self.color, self)
        if value.isValid():
            self.color = value; self._update_color_button()

    def _update_color_button(self) -> None:
        foreground = "#000000" if self.color.lightness() > 150 else "#ffffff"
        self.color_button.setText(self.color.name())
        self.color_button.setStyleSheet(f"background:{self.color.name()};color:{foreground}")

    def selected_label_fields(self) -> list[str]:
        return [
            self.label_columns.item(index).text()
            for index in range(self.label_columns.count())
            if self.label_columns.item(index).checkState() == QtCore.Qt.CheckState.Checked
        ]

    def _validate(self) -> None:
        if not self.name_edit.text().strip():
            QtWidgets.QMessageBox.warning(self, "Import layer", "Enter a layer name."); return
        if self.x_combo.currentText() == self.y_combo.currentText():
            QtWidgets.QMessageBox.warning(self, "Import layer", "Select different X and Y coordinate columns."); return
        self.accept()
