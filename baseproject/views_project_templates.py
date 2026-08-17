import json
from pathlib import Path

import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string

from core.models import UserSettings
from utils.decorators import log_action
from baseproject.utils.project_template_db import ProjectTemplateDB


EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


def _safe_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _load_workbook(file_obj):
    return openpyxl.load_workbook(
        file_obj,
        read_only=True,
        data_only=True,
        keep_vba=False,
    )


def _get_sheet_columns(ws, header_row=1):
    columns = []

    for idx, cell in enumerate(ws[header_row], start=1):
        value = cell.value
        label = str(value).strip() if value is not None else f"Column {idx}"
        columns.append({"index": idx, "label": label})

    return columns


def _parse_sl_groups(request):
    group_nos = request.POST.getlist("sl_group_no[]")
    starts = request.POST.getlist("sl_group_start[]")
    ends = request.POST.getlist("sl_group_end[]")
    directions = request.POST.getlist("sl_group_direction[]")

    groups = []

    for idx in range(max(len(group_nos), len(starts), len(ends))):
        group_no = _safe_int(group_nos[idx] if idx < len(group_nos) else None)
        start_line = _safe_int(starts[idx] if idx < len(starts) else None)
        end_line = _safe_int(ends[idx] if idx < len(ends) else None)
        direction = directions[idx] if idx < len(directions) else "asc"

        if group_no is None and start_line is None and end_line is None:
            continue

        if group_no is None or start_line is None or end_line is None:
            raise ValueError("Each SL group must have group number, start line and end line.")

        if direction not in {"asc", "desc"}:
            direction = "asc"

        groups.append({
            "group_no": group_no,
            "start_line": start_line,
            "end_line": end_line,
            "direction": direction,
        })

    groups.sort(key=lambda g: g["group_no"])

    if not groups:
        raise ValueError("At least one SL group is required.")

    return groups


@login_required
def project_template_excel_sheets(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return JsonResponse({"ok": False, "error": "No file uploaded"}, status=400)

    suffix = Path(file_obj.name).suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        return JsonResponse({"ok": False, "error": "Only .xlsx and .xlsm files are supported"}, status=400)

    try:
        wb = _load_workbook(file_obj)
        return JsonResponse({"ok": True, "sheets": wb.sheetnames})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


@login_required
def project_template_excel_columns(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    file_obj = request.FILES.get("file")
    sheet_name = request.POST.get("sheet", "").strip()

    header_row = _safe_int(request.POST.get("header_row")) or 3
    start_row = _safe_int(request.POST.get("start_row")) or (header_row + 1)

    if not file_obj:
        return JsonResponse({"ok": False, "error": "No file uploaded"}, status=400)

    try:
        wb = _load_workbook(file_obj)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        columns = _get_sheet_columns(ws, header_row=header_row)

        preview = []
        max_preview_row = min(ws.max_row, start_row + 15)

        for row_number in range(max(1, header_row - 2), max_preview_row + 1):
            values = [cell.value for cell in ws[row_number]]
            preview.append({
                "row_number": row_number,
                "values": values,
            })

        return JsonResponse({
            "ok": True,
            "sheet": ws.title,
            "columns": columns,
            "preview": preview,
            "header_row": header_row,
            "start_row": start_row,
        })

    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


@login_required
@log_action("project_template_excel_save", object_type="BASEPROJECT")
def project_template_excel_save(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    user_settings, _ = UserSettings.objects.get_or_create(user=request.user)
    project = user_settings.active_project

    if not project:
        return JsonResponse({"ok": False, "error": "No active project"}, status=400)

    file_obj = request.FILES.get("file")
    sheet_name = request.POST.get("sheet", "").strip()

    if not file_obj:
        return JsonResponse({"ok": False, "error": "No file uploaded"}, status=400)

    suffix = Path(file_obj.name).suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        return JsonResponse({"ok": False, "error": "Only .xlsx and .xlsm files are supported"}, status=400)

    try:
        sl_groups = _parse_sl_groups(request)

        col_first_sl = _safe_int(request.POST.get("col_first_sl"))
        col_last_sl = _safe_int(request.POST.get("col_last_sl"))
        col_lnum = _safe_int(request.POST.get("col_lnum"))
        col_rline = _safe_int(request.POST.get("col_rline"))
        col_tier = _safe_int(request.POST.get("col_tier"))

        default_tier = _safe_int(request.POST.get("default_tier")) or 1
        header_row = _safe_int(request.POST.get("header_row")) or 3
        start_row = _safe_int(request.POST.get("start_row")) or (header_row + 1)

        save_mode = request.POST.get("save_mode", "append").strip().lower()
        replace = save_mode == "replace"

        deployed_by_vessel = _safe_int(request.POST.get("deployed_by_vessel"))
        recovered_by_vessel = _safe_int(request.POST.get("recovered_by_vessel"))

        if any(v is None for v in [col_first_sl, col_last_sl, col_lnum, col_rline]):
            return JsonResponse({
                "ok": False,
                "error": "Please select FirstSL, LastSL, LNum and RLine columns.",
            }, status=400)

        if start_row <= header_row:
            return JsonResponse({
                "ok": False,
                "error": "Data row must be greater than Header row.",
            }, status=400)

        wb = _load_workbook(file_obj)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        rows_to_insert = []
        skipped_rows = 0

        for excel_row_number, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True),
            start=start_row,
        ):
            first_sl = _safe_int(row[col_first_sl - 1]) if len(row) >= col_first_sl else None
            last_sl = _safe_int(row[col_last_sl - 1]) if len(row) >= col_last_sl else None
            lnum = _safe_int(row[col_lnum - 1]) if len(row) >= col_lnum else None
            rline = _safe_int(row[col_rline - 1]) if len(row) >= col_rline else None

            if col_tier:
                tier = _safe_int(row[col_tier - 1]) if len(row) >= col_tier else None
                if tier is None:
                    tier = default_tier
            else:
                tier = default_tier

            if first_sl is None and last_sl is None and lnum is None and rline is None:
                skipped_rows += 1
                continue

            if first_sl is None or last_sl is None or lnum is None or rline is None:
                skipped_rows += 1
                continue

            rows_to_insert.append({
                "FirstSL": first_sl,
                "LastSL": last_sl,
                "LNum": lnum,
                "RLine": rline,
                "Tier": tier,
                "deployed_by_vessel": deployed_by_vessel,
                "recovered_by_vessel": recovered_by_vessel,
            })

        if not rows_to_insert:
            return JsonResponse({"ok": False, "error": "No valid rows found to save."}, status=400)

        ptdb = ProjectTemplateDB(project.db_path)

        inserted = ptdb.insert_rows(rows_to_insert, replace=replace)
        groups_saved = ptdb.save_sl_groups(sl_groups, replace=True)

        request.log_extra = {
            "file": file_obj.name,
            "sheet": ws.title,
            "header_row": header_row,
            "start_row": start_row,
            "save_mode": save_mode,
            "inserted": inserted,
            "skipped_rows": skipped_rows,
            "groups_saved": groups_saved,
        }

        return JsonResponse({
            "ok": True,
            "inserted": inserted,
            "skipped_rows": skipped_rows,
            "groups_saved": groups_saved,
            "replace": replace,
            "sheet": ws.title,
        })

    except KeyError:
        return JsonResponse({"ok": False, "error": f"Sheet not found: {sheet_name}"}, status=400)

    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


@login_required
@require_POST
@log_action("project_template_delete", object_type="BASEPROJECT")
def project_template_delete(request):
    user_settings, _ = UserSettings.objects.get_or_create(user=request.user)
    project = user_settings.active_project

    if not project:
        return JsonResponse({"ok": False, "error": "No active project"}, status=400)
    if not project.can_edit(request.user):
        raise PermissionDenied

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return JsonResponse({"ok": False, "error": "Invalid JSON payload"}, status=400)

    raw_ids = payload.get("ids") if isinstance(payload, dict) else None
    if not isinstance(raw_ids, list) or not raw_ids:
        return JsonResponse({"ok": False, "error": "No template rows selected"}, status=400)

    try:
        ids = sorted({int(value) for value in raw_ids})
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "error": "Invalid template row IDs"}, status=400)

    if any(value <= 0 for value in ids):
        return JsonResponse({"ok": False, "error": "Invalid template row IDs"}, status=400)

    ptdb = ProjectTemplateDB(project.db_path)
    deleted = ptdb.delete_by_ids(ids)
    table_body = ptdb.render_table_body()

    request.log_extra = {"requested_ids": ids, "deleted": deleted}

    return JsonResponse({
        "ok": True,
        "deleted": deleted,
        "table_body": table_body,
    })
