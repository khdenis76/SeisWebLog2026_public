from __future__ import annotations

import csv
import base64
import io
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.template.loader import render_to_string

from core.models import UserSettings
from utils.decorators import log_action

from .forms import StatisticsFilterForm
from .services import ReceiverStatistics


def _active_project(request):
    settings, _ = UserSettings.objects.get_or_create(user=request.user)
    project = settings.active_project
    if not project:
        messages.warning(request, "Select an active project first.")
        return None, redirect("projects")
    if not project.can_view(request.user):
        raise PermissionDenied("You are not a member of this project.")
    if not project.db_path or not Path(project.db_path).exists():
        messages.warning(request, "The active project database was not found.")
        return None, redirect("projects")
    return project, None


def _report(request):
    project, response = _active_project(request)
    if response:
        return None, response
    # An empty query string is a valid "whole database" selection.
    line_choices = ReceiverStatistics.available_lines(project.db_path)
    data = request.GET.copy()
    if not request.GET:
        data.setlist("lines", [str(line) for line in line_choices])
    data.setdefault("time_basis", "either")
    data.setdefault("grouping", "day")
    data.setdefault("period_type", "all")
    form = StatisticsFilterForm(data, line_choices=line_choices)
    if not form.is_valid():
        return form, None
    try:
        payload = ReceiverStatistics(project.db_path, form.cleaned_data).build()
    except (ValueError, OSError) as exc:
        messages.error(request, str(exc))
        payload = None
    return form, (project, payload)


@login_required
@log_action("receiver_statistics", object_type="DSR")
def dashboard(request):
    form, result = _report(request)
    if isinstance(result, HttpResponse):
        return result
    project, payload = result if result else (_active_project(request)[0], None)
    return render(request, "statistics_report/dashboard.html", {
        "form": form, "project": project, "payload": payload,
        "query_string": request.GET.urlencode(),
    })


def _export_payload(request):
    form, result = _report(request)
    if isinstance(result, HttpResponse):
        return None, result
    if not form.is_valid() or not result or not result[1]:
        return None, redirect("statistics_report:dashboard")
    return result, None


@login_required
@log_action("export_receiver_statistics_csv", object_type="DSR")
def export_csv(request):
    result, response = _export_payload(request)
    if response:
        return response
    _, payload = result
    output = io.StringIO()
    fields = ["phase", "rov", "name", "count", "bias_de", "bias_dn", "bias_2d", "inline_mean", "inline_std", "crossline_mean", "crossline_std", "mean", "median", "std", "rms", "cep50", "cep90", "cep95", "cep99", "max", "sigma_e95_mean", "sigma_n95_mean", "within_1", "within_2", "within_5", "within_10", "out_count", "out_pct", "dominant_sector"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(payload["comparisons"])
    response = HttpResponse("\ufeff" + output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="receiver_statistics.csv"'
    return response


@login_required
@log_action("export_receiver_statistics_excel", object_type="DSR")
def export_excel(request):
    result, response = _export_payload(request)
    if response:
        return response
    _, payload = result
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["TGS — SeisWebLog Receiver Statistics"])
    summary.append(["Project", payload["project_name"]])
    summary.append([])
    summary.append(["Metric", "Value"])
    for key, value in payload["summary"].items():
        summary.append([key.replace("_", " ").title(), value])
    summary.append([])
    summary.append(["Phase", "ROV", "Comparison", "N", "Bias DE", "Bias DN", "2D Bias", "In-line mean", "In-line STD", "Cross-line mean", "Cross-line STD", "Radial mean", "CEP50", "CEP95", "STD", "RMS", "Max", "Sigma E 95% mean", "Sigma N 95% mean", "Out of spec", "Out %", "Dominant sector"])
    for row in payload["comparisons"]:
        summary.append([row[k] for k in ("phase", "rov", "name", "count", "bias_de", "bias_dn", "bias_2d", "inline_mean", "inline_std", "crossline_mean", "crossline_std", "mean", "cep50", "cep95", "std", "rms", "max", "sigma_e95_mean", "sigma_n95_mean", "out_count", "out_pct", "dominant_sector")])

    grouped = workbook.create_sheet("Grouped Activity")
    grouped.append(["Period / group", "Deployed", "Recovered"])
    for row in payload["grouped"]:
        grouped.append([row["group"], row["Deployed"], row["Recovered"]])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="005A8B")
            cell.alignment = Alignment(vertical="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(28, max(10, max(len(str(c.value or "")) for c in column) + 2))
    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="receiver_statistics.xlsx"'
    return response


def _logo_uri():
    logo = Path(__file__).resolve().parent.parent / "logos" / "2024_TGS_logo_blue.png"
    if not logo.exists():
        return ""
    return "data:image/png;base64," + base64.b64encode(logo.read_bytes()).decode("ascii")


@login_required
@log_action("export_receiver_statistics_pdf", object_type="DSR")
def export_pdf(request):
    result, response = _export_payload(request)
    if response:
        return response
    project, payload = result
    from .latex_export import render_pdf
    logo = Path(__file__).resolve().parent.parent / "logos" / "2024_TGS_logo_blue.png"
    try:
        pdf = render_pdf(payload, logo)
    except RuntimeError as exc:
        return HttpResponse(str(exc), status=500, content_type="text/plain; charset=utf-8")
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="receiver_statistics.pdf"'
    return response


@login_required
@log_action("export_receiver_statistics_html", object_type="DSR")
def export_html(request):
    result, response = _export_payload(request)
    if response:
        return response
    project, payload = result
    from plotly.offline import get_plotlyjs
    html = render_to_string("statistics_report/export_html.html", {
        "project": project, "payload": payload, "logo_uri": _logo_uri(), "plotly_js": get_plotlyjs(),
    })
    response = HttpResponse(html, content_type="text/html; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="receiver_statistics_interactive.html"'
    return response
